"""Módulo de extracción de movimientos BBVA.

Este módulo expone funciones para leer un estado de cuenta PDF y generar un
archivo Excel con los movimientos.

El módulo intenta leer texto directamente del PDF usando pdfplumber. Si la página
está escaneada (imagen), usa EasyOCR para realizar OCR y reconstruir la tabla.

El código está pensado para ser usado desde una GUI (tkinter) u otro front-end.
"""

import os
import re
import unicodedata
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

import certifi
import numpy as np
import pdfplumber

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# Tipos
Movimiento = Dict[str, Optional[str]]
InfoCuenta = Dict[str, Optional[str]]
LogFunc = Callable[[str], None]

# Patrones de reconocimiento
PATRON_FECHA = re.compile(r"(\d{1,2}/[A-Za-z]{3})")
PATRON_MONTO = re.compile(r"(\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2}))")


def _norm_text(s: str) -> str:
    """Normaliza texto para comparaciones (sin acentos, mayúsculas)."""
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r'[^A-Za-z0-9 ]', '', s)
    return s.strip().upper()


def _normalizar_monto(texto: Optional[str]) -> Optional[float]:
    if texto is None:
        return None
    s = str(texto).strip()
    if not s:
        return None
    s = s.replace('$', '').replace(' ', '').replace('US$', '')
    # Convertir comas a punto si solo hay comas
    if s.count(',') == 1 and s.count('.') == 0:
        s = s.replace(',', '.')
    else:
        s = s.replace(',', '')
    try:
        return float(s)
    except Exception:
        return None


def _init_easyocr_reader(log: LogFunc):
    """Inicializa y devuelve un reader de EasyOCR."""
    try:
        import easyocr
    except ImportError:
        log('⚠ easyocr no instalado; no se puede hacer OCR en PDF escaneados.')
        return None

    os.environ.setdefault('SSL_CERT_FILE', certifi.where())

    try:
        log('🔎 Inicializando EasyOCR...')
        reader = easyocr.Reader(['es'], gpu=False)
        return reader
    except Exception as e:
        log(f'⚠ Error al inicializar EasyOCR: {e}')
        return None


def _texto_desde_ocr(reader, pagina, log: LogFunc) -> Tuple[str, List[Tuple[Tuple[float, float, float, float], str]]]:
    """Extrae texto de la página con EasyOCR y devuelve texto y cajas."""
    img = pagina.to_image(resolution=200).original
    arr = np.array(img)
    log('🔎 Ejecutando OCR (EasyOCR)...')
    resultados = reader.readtext(arr)
    texto = '\n'.join([t for _, t, _ in resultados])
    return texto, resultados


def _parsear_encabezado(texto: str) -> InfoCuenta:
    info: InfoCuenta = {
        'numero_cuenta': None,
        'periodo': None,
        'fecha_corte': None,
        'saldo_anterior': None,
        'saldo_final': None,
        'total_cargos': None,
        'total_abonos': None,
    }

    # Buscar número de cuenta
    match = re.search(r'CUENTA\s*[:\-]?\s*(\d{4,})', texto, re.IGNORECASE)
    if match:
        info['numero_cuenta'] = match.group(1).strip()

    # Período
    match = re.search(r'PERI[ÓO]DO\s*[:\-]?\s*([\d/\-\sA-Za-z]+)', texto, re.IGNORECASE)
    if match:
        info['periodo'] = match.group(1).strip()

    # Fecha de corte
    match = re.search(r'CORTE\s*[:\-]?\s*(\d{2}/\d{2}/\d{4})', texto, re.IGNORECASE)
    if match:
        info['fecha_corte'] = match.group(1).strip()

    # Saldos y totales
    match = re.search(r'SALDO\s+ANTERIOR\s*[:\-]?\s*\$?\s*([\d.,]+)', texto, re.IGNORECASE)
    if match:
        info['saldo_anterior'] = match.group(1).strip()

    match = re.search(r'SALDO\s+FINAL\s*[:\-]?\s*\$?\s*([\d.,]+)', texto, re.IGNORECASE)
    if match:
        info['saldo_final'] = match.group(1).strip()

    match = re.search(r'TOTAL\s+CARGOS\s*[:\-]?\s*\$?\s*([\d.,]+)', texto, re.IGNORECASE)
    if match:
        info['total_cargos'] = match.group(1).strip()

    match = re.search(r'TOTAL\s+ABONOS\s*[:\-]?\s*\$?\s*([\d.,]+)', texto, re.IGNORECASE)
    if match:
        info['total_abonos'] = match.group(1).strip()

    return info


def _parsear_movimiento_linea(linea: str) -> Optional[Movimiento]:
    """Extrae un movimiento de una línea de texto."""
    linea = linea.strip()
    if not linea:
        return None

    m_fecha = PATRON_FECHA.match(linea)
    if not m_fecha:
        return None

    fecha_oper = m_fecha.group(1)
    contenido = linea[m_fecha.end():].strip()

    # Intento de separación por columnas basadas en varios espacios
    partes = re.split(r"\s{2,}", contenido)

    descripcion = partes[0] if partes else ''
    referencia = partes[1] if len(partes) > 1 else ''
    montos_texto = ' '.join(partes[2:]) if len(partes) > 2 else ''

    montos = PATRON_MONTO.findall(montos_texto)

    cargo = None
    abono = None
    if montos:
        if len(montos) >= 2:
            cargo = _normalizar_monto(montos[0])
            abono = _normalizar_monto(montos[1])
        else:
            texto_lower = linea.lower()
            if 'cargo' in texto_lower and 'abono' not in texto_lower:
                cargo = _normalizar_monto(montos[0])
            elif 'abono' in texto_lower and 'cargo' not in texto_lower:
                abono = _normalizar_monto(montos[0])
            else:
                abono = _normalizar_monto(montos[0])

    tipo = 'CARGO' if cargo else 'ABONO'

    return {
        'Fecha_Oper': fecha_oper,
        'Fecha_Liq': fecha_oper,
        'Descripcion': descripcion,
        'Referencia': referencia,
        'Cargo': cargo,
        'Abono': abono,
        'Saldo_Oper': None,
        'Saldo_Liq': None,
        'Tipo': tipo,
    }


def _asignar_columna_por_x(x: float, header_positions: Dict[str, float]) -> Optional[str]:
    """Encuentra la columna (header) más cercana en X para un texto dado."""
    if not header_positions:
        return None
    mejor = min(header_positions.items(), key=lambda kv: abs(kv[1] - x))
    return mejor[0]


def _parsear_movimientos_desde_ocr(
    resultados: List[Tuple[Tuple[float, float, float, float], str, float]],
    log: LogFunc,
) -> List[Movimiento]:
    """Construye movimientos a partir de resultados de EasyOCR."""
    # Identificar encabezados y sus posiciones
    header_positions: Dict[str, float] = {}
    for bbox, texto, _ in resultados:
        norm = _norm_text(texto)
        if norm in ('FECHA', 'OPER', 'LIQ', 'LIQUIDACION', 'DESCRIPCION', 'REFERENCIA', 'CARGOS', 'ABONOS', 'OPERACION'):
            # Guardar la posición X del encabezado
            header_positions[norm] = bbox[0][0]

    # Desechamos encabezados y construimos filas por posición Y
    filas: Dict[int, Dict[str, str]] = {}

    # Determinar y de inicio de tabla (después del header)
    y_min_header = min((bbox[0][1] for bbox, texto, _ in resultados if _norm_text(texto) in ('FECHA', 'OPER', 'LIQ', 'DESCRIPCION')), default=0)

    for bbox, texto, _ in resultados:
        y_center = (bbox[0][1] + bbox[2][1]) / 2
        if y_center <= y_min_header + 5:
            continue

        x_center = (bbox[0][0] + bbox[1][0]) / 2
        col = _asignar_columna_por_x(x_center, header_positions)
        if not col:
            continue

        # Agrupar por línea (redondeo de Y)
        y_key = int(round(y_center / 10) * 10)
        fila = filas.setdefault(y_key, {})
        fila[col] = (fila.get(col, '') + ' ' + texto).strip()

    movimientos: List[Movimiento] = []

    for y in sorted(filas.keys()):
        fila = filas[y]
        if not fila:
            continue

        fecha_oper = fila.get('FECHA') or fila.get('OPER')
        fecha_liq = fila.get('LIQ')
        descripcion = fila.get('DESCRIPCION', '')
        referencia = fila.get('REFERENCIA', '')

        cargo = _normalizar_monto(fila.get('CARGOS'))
        abono = _normalizar_monto(fila.get('ABONOS'))
        saldo_oper = _normalizar_monto(fila.get('OPERACION'))
        saldo_liq = _normalizar_monto(fila.get('LIQUIDACION'))

        # Fallbacks sencillos
        if not fecha_liq:
            fecha_liq = fecha_oper

        tipo = 'CARGO' if cargo else 'ABONO'

        if not fecha_oper and not descripcion:
            continue

        movimientos.append(
            {
                'Fecha_Oper': fecha_oper,
                'Fecha_Liq': fecha_liq,
                'Descripcion': descripcion,
                'Referencia': referencia,
                'Cargo': cargo,
                'Abono': abono,
                'Saldo_Oper': saldo_oper,
                'Saldo_Liq': saldo_liq,
                'Tipo': tipo,
            }
        )

    return movimientos


def extraer_movimientos_desde_pdf(
    pdf_path: str, log: LogFunc = print
) -> Tuple[List[Movimiento], InfoCuenta]:
    """Extrae movimientos y encabezado desde un PDF."""
    movimientos: List[Movimiento] = []
    info_cuenta: InfoCuenta = {}

    ruta = Path(pdf_path)
    if not ruta.exists():
        raise FileNotFoundError(f"No se encontró el archivo PDF: {pdf_path}")

    log(f"📄 Abriendo PDF: {ruta.name}")
    reader = _init_easyocr_reader(log)

    with pdfplumber.open(ruta) as pdf:
        for idx, pagina in enumerate(pdf.pages, start=1):
            log(f"  Página {idx}/{len(pdf.pages)}")

            texto = pagina.extract_text() or ''
            if not texto.strip() and reader:
                texto, resultados = _texto_desde_ocr(reader, pagina, log)
                if not texto.strip():
                    log("    No se pudo extraer texto en esta página (OCR falló).")
                    continue

                if idx == 1:
                    info_cuenta = _parsear_encabezado(texto)
                    log("    Encabezado parseado.")

                movs = _parsear_movimientos_desde_ocr(resultados, log)
                movimientos.extend(movs)
                continue

            if idx == 1:
                info_cuenta = _parsear_encabezado(texto)
                log("    Encabezado parseado.")

            # Intentar parsear movimientos a partir del texto plano
            lineas = [l.strip() for l in texto.splitlines() if l.strip()]
            for linea in lineas:
                mov = _parsear_movimiento_linea(linea)
                if mov:
                    movimientos.append(mov)

    log(f"✓ Movimientos encontrados: {len(movimientos)}")
    return movimientos, info_cuenta


def generar_excel_movimientos(
    movimientos: List[Movimiento],
    info_cuenta: InfoCuenta,
    ruta_salida: str,
    log: LogFunc = print,
) -> None:
    """Genera el archivo Excel con dos hojas: Movimientos y Resumen."""

    ruta = Path(ruta_salida)
    ruta.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    encabezados = [
        'Fecha_Oper',
        'Fecha_Liq',
        'Descripción',
        'Referencia',
        'Cargo',
        'Abono',
        'Saldo_Oper',
        'Saldo_Liq',
        'Tipo',
    ]
    ws.append(encabezados)

    header_fill = PatternFill(start_color='003F87', end_color='003F87', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for idx, mov in enumerate(movimientos, start=2):
        ws.append([
            mov.get('Fecha_Oper'),
            mov.get('Fecha_Liq'),
            mov.get('Descripcion'),
            mov.get('Referencia'),
            mov.get('Cargo'),
            mov.get('Abono'),
            mov.get('Saldo_Oper'),
            mov.get('Saldo_Liq'),
            mov.get('Tipo'),
        ])

        if idx % 2 == 0:
            fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            for cell in ws[idx]:
                cell.fill = fill

    total_cargos = sum(m.get('Cargo') or 0 for m in movimientos)
    total_abonos = sum(m.get('Abono') or 0 for m in movimientos)

    total_row = len(movimientos) + 3
    ws[f'A{total_row}'] = 'TOTALES'
    ws[f'E{total_row}'] = total_cargos
    ws[f'F{total_row}'] = total_abonos

    ws[f'E{total_row}'].number_format = '0.00'
    ws[f'F{total_row}'].number_format = '0.00'
    ws[f'A{total_row}'].font = Font(bold=True)

    # Autoajustar columnas
    anchos = [12, 12, 40, 18, 15, 15, 15, 15, 12]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # Hoja resumen
    ws2 = wb.create_sheet('Resumen')
    ws2['A1'] = 'RESUMEN DE CUENTA'
    ws2['A1'].font = Font(bold=True, size=14)

    ws2['A3'] = 'Número de Movimientos:'
    ws2['B3'] = len(movimientos)

    ws2['A4'] = 'Total Cargos:'
    ws2['B4'] = total_cargos
    ws2['B4'].number_format = '0.00'

    ws2['A5'] = 'Total Abonos:'
    ws2['B5'] = total_abonos
    ws2['B5'].number_format = '0.00'

    ws2['A7'] = 'Concepto'
    ws2['B7'] = 'Monto'
    ws2['A8'] = 'Cargos'
    ws2['B8'] = total_cargos
    ws2['A9'] = 'Abonos'
    ws2['B9'] = total_abonos

    chart = BarChart()
    chart.title = 'Cargos vs Abonos'

    # Usar Reference para la serie de datos
    data = Reference(ws2, min_col=2, min_row=7, max_row=9)
    chart.add_data(data, titles_from_data=True)
    ws2.add_chart(chart, 'D3')

    log(f"💾 Guardando Excel en: {ruta}")
    wb.save(str(ruta))


def _formatear_log(log_func: LogFunc, mensaje: str):
    log_func(mensaje)


if __name__ == '__main__':
    # Prueba rápida en consola
    def _log(m):
        print(m)

    pdf = 'doc01266120250626153613.pdf'
    excel = 'movimientos_bbva.xlsx'
    movs, info = extraer_movimientos_desde_pdf(pdf, log=_log)
    generar_excel_movimientos(movs, info, excel, log=_log)
